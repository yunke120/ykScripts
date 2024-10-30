import os
import shutil
import tkinter as tk
from tkinter import filedialog, ttk
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import webbrowser

def extract_table_column(doc_path, column_name):
    doc = Document(doc_path)
    results = []
    for table in doc.tables:
        header_row = table.rows[0]
        column_index = None
        for i, cell in enumerate(header_row.cells):
            if cell.text.strip().lower() == column_name.lower():
                column_index = i
                break
        if column_index is not None:
            for row in table.rows[1:]:
                cell_value = row.cells[column_index].text.strip()
                if cell_value and cell_value.lower() != column_name.lower():
                    results.append(cell_value)
    return results

def find_documents(folder_path):
    document_files = []
    allowed_extensions = ('.pdf', '.doc', '.docx')
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(allowed_extensions):
                full_path = os.path.join(root, file)
                file_name = os.path.splitext(file)[0]
                document_files.append((file_name, full_path))
    return document_files

def fuzzy_match(spec, file_name, fuzzy_level):
    return spec[:-fuzzy_level].lower() in file_name.lower()

def smart_match(spec, file_name):
    if spec.lower() in file_name.lower():
        return 0  # 精确匹配
    for level in range(1, 4):
        if spec[:-level].lower() in file_name.lower():
            return level  # 返回模糊匹配的级别
    return -1  # 未匹配

def write_to_excel_and_copy_files(documents, specifications, output_file, exists_folder, not_exist_file, search_mode, fuzzy_level=1):
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    # 写入表头
    ws.append(["序号", "文件名", "文件绝对地址", "匹配类型", "匹配的规格型号"])

    # 创建not_exist.xlsx
    not_exist_wb = Workbook()
    not_exist_ws = not_exist_wb.active
    not_exist_ws.append(["序号", "型号规格"])

    # 定义填充颜色
    fill_colors = {
        0: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # 绿色 - 精确匹配
        1: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # 黄色 - 一级模糊
        2: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # 橙色 - 二级模糊
        3: PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid"),  # 粉色 - 三级模糊
    }

    found_specs = set()
    remaining_specs = set(specifications)
    spec_folders = {}  # 用于存储每个规格型号对应的文件夹名称

    # 用于给文件夹添加序号的计数器
    folder_counter = 1

    # 首先将所有文档写入Excel
    for idx, (file_name, full_path) in enumerate(documents, start=1):
        row = [idx, file_name, full_path, "", ""]
        ws.append(row)

    if search_mode == "smart":
        # 智能匹配逻辑
        for match_level in range(4):  # 0: 精确匹配, 1-3: 模糊匹配级别
            for idx, (file_name, full_path) in enumerate(documents, start=1):
                matched_specs = set()
                for spec in remaining_specs:
                    if match_level == 0:
                        if spec.lower() in file_name.lower():
                            matched_specs.add(spec)
                    else:
                        if spec[:-match_level].lower() in file_name.lower():
                            matched_specs.add(spec)
                
                if matched_specs:
                    for cell in ws[idx + 1]:  # +1 因为第一行是表头
                        cell.fill = fill_colors[match_level]
                    ws.cell(row=idx + 1, column=4, value=f"{'精确' if match_level == 0 else f'{match_level}级模糊'}")
                    ws.cell(row=idx + 1, column=5, value=", ".join(matched_specs))
                    found_specs.update(matched_specs)
                    
                    # 为每个匹配的规格型号创建文件夹并复制文件
                    for spec in matched_specs:
                        if spec not in spec_folders:
                            safe_spec = spec.replace('/', '-')
                            safe_spec = "".join([c for c in safe_spec if c.isalnum() or c in (' ', '-', '_')])
                            folder_name = f"{folder_counter:03d}-{safe_spec}"
                            spec_folders[spec] = folder_name
                            folder_counter += 1
                        else:
                            folder_name = spec_folders[spec]
                        
                        spec_folder = os.path.join(exists_folder, folder_name)
                        os.makedirs(spec_folder, exist_ok=True)
                        dest_path = os.path.join(spec_folder, os.path.basename(full_path))
                        shutil.copy2(full_path, dest_path)
                        print(f"已复制文件: {os.path.basename(full_path)} 到 {folder_name} 文件夹")
            
            remaining_specs -= found_specs
            if not remaining_specs:
                break
    elif search_mode == "fuzzy":
        # 模糊匹配逻辑
        for idx, (file_name, full_path) in enumerate(documents, start=1):
            for spec in specifications:
                if fuzzy_match(spec, file_name, fuzzy_level):
                    for cell in ws[idx + 1]:
                        cell.fill = fill_colors[fuzzy_level]
                    ws.cell(row=idx + 1, column=4, value=f"{fuzzy_level}级模糊")
                    ws.cell(row=idx + 1, column=5, value=spec)
                    found_specs.add(spec)
                    
                    # 创建规格型号文件夹并复制文件
                    if spec not in spec_folders:
                        safe_spec = spec.replace('/', '-')
                        safe_spec = "".join([c for c in safe_spec if c.isalnum() or c in (' ', '-', '_')])
                        folder_name = f"{folder_counter:03d}-{safe_spec}"
                        spec_folders[spec] = folder_name
                        folder_counter += 1
                    else:
                        folder_name = spec_folders[spec]
                    
                    spec_folder = os.path.join(exists_folder, folder_name)
                    os.makedirs(spec_folder, exist_ok=True)
                    dest_path = os.path.join(spec_folder, os.path.basename(full_path))
                    shutil.copy2(full_path, dest_path)
                    print(f"已复制文件: {os.path.basename(full_path)} 到 {folder_name} 文件夹")
                    break  # 一旦找到匹配，就跳出内层循环
    else:  # 精确查询
        for idx, (file_name, full_path) in enumerate(documents, start=1):
            for spec in specifications:
                if spec.lower() in file_name.lower():
                    for cell in ws[idx + 1]:
                        cell.fill = fill_colors[0]
                    ws.cell(row=idx + 1, column=4, value="精确")
                    ws.cell(row=idx + 1, column=5, value=spec)
                    found_specs.add(spec)
                    
                    # 创建规格型号文件夹并复制文件
                    if spec not in spec_folders:
                        safe_spec = spec.replace('/', '-')
                        safe_spec = "".join([c for c in safe_spec if c.isalnum() or c in (' ', '-', '_')])
                        folder_name = f"{folder_counter:03d}-{safe_spec}"
                        spec_folders[spec] = folder_name
                        folder_counter += 1
                    else:
                        folder_name = spec_folders[spec]
                    
                    spec_folder = os.path.join(exists_folder, folder_name)
                    os.makedirs(spec_folder, exist_ok=True)
                    dest_path = os.path.join(spec_folder, os.path.basename(full_path))
                    shutil.copy2(full_path, dest_path)
                    print(f"已复制文件: {os.path.basename(full_path)} 到 {folder_name} 文件夹")
                    break  # 一旦找到匹配，就跳出内层循环

    # 保存Excel文件
    wb.save(output_file)
        
    # 写入未找到的型号规格到not_exist.xlsx，并创建空文件夹
    not_found_specs = set(specifications) - found_specs
    for idx, spec in enumerate(not_found_specs, start=1):
        not_exist_ws.append([idx, spec])
        safe_spec = spec.replace('/', '-')
        safe_spec = "".join([c for c in safe_spec if c.isalnum() or c in (' ', '-', '_')])
        folder_name = f"{folder_counter:03d}-{safe_spec}（空）"
        empty_folder = os.path.join(exists_folder, folder_name)
        os.makedirs(empty_folder, exist_ok=True)
        print(f"已创建空文件夹: {folder_name}")
        folder_counter += 1

    not_exist_wb.save(not_exist_file)

    return found_specs, not_found_specs

class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("元器件手册定位查询软件 V1.0")
        self.geometry("600x500")

        self.word_path = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.exists_folder = tk.StringVar()
        self.search_mode = tk.StringVar(value="exact")
        self.fuzzy_level = tk.IntVar(value=1)

        self.create_widgets()

    def create_widgets(self):
        # Word文档选择
        ttk.Label(self, text="待处理Word:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(self, textvariable=self.word_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self, text="浏览", command=self.select_word_file).grid(row=0, column=2, padx=5, pady=5)

        # 元器件手册文件夹
        ttk.Label(self, text="元器件手册:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(self, textvariable=self.folder_path, width=50).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(self, text="浏览", command=self.select_folder).grid(row=1, column=2, padx=5, pady=5)

        # 保存文件夹
        ttk.Label(self, text="保存文件夹:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(self, textvariable=self.exists_folder, width=50).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(self, text="浏览", command=self.select_exists_folder).grid(row=2, column=2, padx=5, pady=5)

        # 查询模式选择
        ttk.Label(self, text="查询模式:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        ttk.Radiobutton(self, text="精确查询", variable=self.search_mode, value="exact", command=self.update_fuzzy_state).grid(row=3, column=1, sticky="w", padx=5, pady=5)
        ttk.Radiobutton(self, text="模糊查询", variable=self.search_mode, value="fuzzy", command=self.update_fuzzy_state).grid(row=3, column=1, padx=5, pady=5)
        ttk.Radiobutton(self, text="智能查询", variable=self.search_mode, value="smart", command=self.update_fuzzy_state).grid(row=3, column=1, sticky="e", padx=5, pady=5)

        # 模糊查询等级选择
        ttk.Label(self, text="模糊查询等级:").grid(row=4, column=0, sticky="w", padx=5, pady=5)
        self.level_frame = ttk.Frame(self)
        self.level_frame.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        self.fuzzy_buttons = [
            ttk.Radiobutton(self.level_frame, text="一级", variable=self.fuzzy_level, value=1),
            ttk.Radiobutton(self.level_frame, text="二级", variable=self.fuzzy_level, value=2),
            ttk.Radiobutton(self.level_frame, text="三级", variable=self.fuzzy_level, value=3)
        ]
        for button in self.fuzzy_buttons:
            button.pack(side=tk.LEFT)

        # 初始化模糊查询等级的状态
        self.update_fuzzy_state()

        ttk.Button(self, text="帮助", command=self.open_help).grid(row=5, column=0, padx=5, pady=5)
        # 输出按钮
        # _style = ttk.Style()
        # _style.configure('TButton', foreground="black", background="green") #  style='TButton', 
        ttk.Button(self, text="输出",command=self.process_documents).grid(row=5, column=2, padx=5, pady=5)

        # 日志显示
        self.log_text = tk.Text(self, height=15, width=70)
        self.log_text.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

        # 滚动条
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=6, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def open_help(self):
        help_file = os.path.join(os.path.dirname(__file__), "help.html")
        if os.path.exists(help_file):
            webbrowser.open('file://' + os.path.realpath(help_file))
        else:
            self.log("帮助文件不存在。请确保 'help.html' 文件在程序同一目录下。")
    def select_word_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Word Document", "*.docx")])
        if filename:
            self.word_path.set(filename)

    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_path.set(folder)

    def select_exists_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.exists_folder.set(folder)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.update_idletasks()

    def process_documents(self):
        word_doc_path = self.word_path.get()
        folder_path = self.folder_path.get()
        output_file = "元器件清单.xlsx"
        not_exist_file = "not_exist.xlsx"
        exists_folder = self.exists_folder.get()

        if not word_doc_path or not folder_path or not exists_folder:
            self.log("请填写所有必要的信息。")
            return

        self.log("开始处理...")

        # 从Word文档中提取型号规格
        specifications = extract_table_column(word_doc_path, '型号规格')
        self.log(f"从Word文档中提取了 {len(specifications)} 个型号规格")

        # 查找文档
        documents = find_documents(folder_path)
        self.log(f"在文件夹中找到 {len(documents)} 个文档")

        # 写入Excel，检查型号规格，并复制匹配的文件
        found_specs, not_found_specs = write_to_excel_and_copy_files(
            documents, specifications, output_file, exists_folder, not_exist_file, 
            search_mode=self.search_mode.get(), fuzzy_level=self.fuzzy_level.get()
        )

        self.log(f"\n文档信息已写入 {output_file}")
        self.log(f"总共找到 {len(documents)} 个文档")
        self.log(f"匹配的文件已复制到 {os.path.abspath(exists_folder)} 文件夹")
        self.log(f"\n已找到的型号规格数量: {len(found_specs)}")
        self.log(f"未找到的型号规格数量: {len(not_found_specs)}")

        if not_found_specs:
            self.log(f"\n未找到的型号规格已写入 {not_exist_file}")
        else:
            self.log("\n所有型号规格都在文件名中找到了。")

        self.log("处理完成。")

    def update_fuzzy_state(self):
        state = 'normal' if self.search_mode.get() == 'fuzzy' else 'disabled'
        for button in self.fuzzy_buttons:
            button.configure(state=state)

if __name__ == "__main__":
    app = App()
    app.mainloop()

